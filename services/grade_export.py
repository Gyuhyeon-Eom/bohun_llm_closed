# -*- coding: utf-8 -*-
"""상이등급 심사표 xlsx 산출물 (가로형 — 원본 양식).
업로드된 실제 양식처럼 컬럼을 좌→우로 나열하고 안건을 행으로 표시.
단건(ga_id) 또는 전체(ga_id=None) export 지원. 작업로그는 별도 시트.
"""
import os
import tempfile


def _fetch(ga_id):
    import psycopg
    from psycopg.rows import dict_row
    from config.settings import PG_DSN
    with psycopg.connect(PG_DSN, row_factory=dict_row) as conn, conn.cursor() as cur:
        if ga_id:
            cur.execute("SELECT * FROM grade_agenda WHERE ga_id=%s", (ga_id,))
            ags = [cur.fetchone()]
        else:
            cur.execute("SELECT * FROM grade_agenda ORDER BY ga_id")
            ags = cur.fetchall()
        ags = [a for a in ags if a]
        logs_by = {}
        if ga_id and ags:
            cur.execute("SELECT step, event, actor, detail, file_name, status, created_at"
                        " FROM grade_log WHERE ga_id=%s ORDER BY gl_id", (ags[0]["ga_id"],))
            logs_by[ags[0]["ga_id"]] = cur.fetchall()
    return ags, logs_by


# 심사표 컬럼 (좌 → 우) — 확정 양식 14컬럼 (260720, 실물 양식 사진 대조).
# scope: "ag"=안건(인적사항) 단위 — 상이처가 여러 개면 세로 병합 / "item"=상이처별 행.
# 로직(한 플로우): 요건인정 상이처 → 직전등급 → 신검과목 → 신검등급 → 상이정도 및 소견
#   → 관련자료 → 검토사항 → 비고 → 상이처별 제안등급이 전부 그 상이처의 행에 붙고,
#   마지막에 상이처별 제안등급 중 최고(중한) 등급으로 종합 제안등급만 안건 단위 산출.
COLUMNS = [
    ("신검종류", "apply_type", 11, "ag"),
    ("성명", "applicant", 9, "ag"),
    ("주민등록번호", "resident_no", 15, "ag"),
    ("대상구분", "target_type", 11, "ag"),
    ("요건인정 상이처", "injury", 20, "item"),
    ("직전등급", "prev_grade", 14, "item"),
    ("신검과목", "exam_dept", 11, "item"),
    ("신검등급", "exam_grade", 12, "item"),
    ("상이정도 및 소견\n(보훈병원 전문의)", "__opinion__", 44, "item"),
    ("관련자료", "related_docs", 20, "item"),
    ("검토사항", "review_items", 40, "item"),
    ("비고", "note_items", 20, "item"),
    ("상이처별 제안등급", "__grade_each__", 16, "item"),
    ("종합 제안등급", "__grade_total__", 14, "ag"),
]


def _prev_grade(ag):
    """직전등급: grade_change('7급 8122호 → 재심의 대상')의 화살표 앞부분."""
    gc = ag.get("grade_change") or ""
    return gc.split("→")[0].strip() or "—"


def _items(ag):
    """상이처별 항목 리스트. injury_items(JSONB) 우선, 없으면 기존 단일 컬럼으로 1건 구성."""
    import json
    items = ag.get("injury_items")
    if isinstance(items, str):
        try:
            items = json.loads(items)
        except Exception:
            items = None
    if not items:
        items = [{"injury": ag.get("yeu_injury") or ag.get("injury"),
                  "body_part": ag.get("body_part"), "prev_grade": _prev_grade(ag),
                  "exam_dept": ag.get("exam_dept"), "exam_grade": ag.get("exam_grade"),
                  "opinion": None}]
    # 상이처별 검토사항·비고·관련자료 — 항목에 없으면 대표(첫) 상이처가 안건 값을 승계
    for k, it in enumerate(items):
        for key in ("review_items", "note_items", "related_docs"):
            if it.get(key) is None and k == 0:
                it[key] = ag.get(key)
    return items


def _predict_grade(injury, body_part, emb):
    """상이처별 AI 제안등급 — grade_predict(별표3 대조)로 산출. 실패 시 None."""
    if emb is None or not injury:
        return None
    try:
        from services import grade_predict
        r = grade_predict.predict(injury, body_part, emb, n=3)
        return r.get("grade1")
    except Exception:
        return None


def _severity(grade):
    """등급 문자열의 중증도 (숫자 작을수록 중함). 미달·불명은 99."""
    import re
    m = re.search(r"(\d)\s*급", grade or "") or re.match(r"\s*(\d)\s*-", grade or "")
    return int(m.group(1)) if m else 99


def _proposed(it, pred):
    """상이처별 제안등급 — 신검등급(신체검사 실측 판정) 기반. 실물 양식과 동일하게
    신검등급을 제안값으로 쓰고, AI 별표3 대조 예측은 참고로만 병기한다
    (예측이 상병명 텍스트만 대조하므로 실측과 동떨어진 등급이 나올 수 있음)."""
    eg = (it.get("exam_grade") or "").strip()
    if eg and eg != "—":
        return eg
    return pred  # 신검등급 미기재 시에만 AI 예측으로 대체


def _updown(prev, proposed):
    """직전등급 대비 승급/하향/유지 표기 (실물 양식의 비고 '승급' 표기 대응)."""
    a, b = _severity(prev), _severity(proposed)
    if b == 99:
        return "등급기준미달" if a < 99 else None
    if a == 99:
        return "승급(기준미달→등급)"
    return "승급" if b < a else ("하향" if b > a else "직전등급 유지")


def _total_grade(preds):
    """종합 제안등급 = 상이처별 제안등급 중 가장 중한(숫자 낮은) 등급."""
    real = [p for p in preds if p and _severity(p) < 99]
    return min(real, key=_severity) if real else None


def export_xlsx(ga_id=None, emb=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    ags, logs_by = _fetch(ga_id)
    if not ags:
        wb = Workbook(); ws = wb.active; ws["A1"] = "안건을 찾을 수 없습니다."
        path = os.path.join(tempfile.gettempdir(), f"grade_{ga_id or 'all'}.xlsx"); wb.save(path)
        return f"grade_{ga_id or 'all'}.xlsx", path

    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="D6E4D2")
    title_font = Font(name="맑은 고딕", size=14, bold=True)
    hf = Font(name="맑은 고딕", size=9.5, bold=True)
    cf = Font(name="맑은 고딕", size=9.5)
    wrap = Alignment(wrap_text=True, vertical="center", horizontal="center")
    wrap_l = Alignment(wrap_text=True, vertical="center", horizontal="left")
    center = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "상이등급심사표"
    ncol = len(COLUMNS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws.cell(1, 1, "상 이 등 급 심 사 표").font = title_font
    ws.cell(1, 1).alignment = center
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    if any(a.get("is_real") for a in ags):  # 실데이터 포함 시 파란 강조 (표본 문구 대신)
        ws.cell(2, 1, "※ 실데이터 포함 — 개인정보 마스킹 적용분, 취급 주의").font = \
            Font(name="맑은 고딕", size=9, bold=True, color="1D4ED8")
    else:
        ws.cell(2, 1, "※ 시연용 표본 — 실제 사례 아님(개인정보 미포함)").font = Font(name="맑은 고딕", size=9, color="B45309")
    ws.cell(2, 1).alignment = center

    hr = 4
    for ci, (label, _, width, _s) in enumerate(COLUMNS, 1):
        c = ws.cell(hr, ci, label)
        c.font = hf; c.fill = hdr_fill; c.border = border
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[hr].height = 40

    def fmt(v):
        if v is None:
            return "—"
        if isinstance(v, (list, tuple)):
            return "\n".join(f"○ {x}" for x in v) if v else "—"
        return str(v)

    def _measure_str(ag):
        import json as _j
        m = ag.get("measurements")
        if not m:
            return ""
        if isinstance(m, str):
            try: m = _j.loads(m)
            except Exception: return ""
        return " / ".join(f"{x.get('name')} {x.get('value')}{x.get('unit','')}"
                          + (f"({x.get('result')})" if x.get('result') else "") for x in m)

    def soken(ag):
        """상이정도 및 소견(보훈병원 전문의) — 측정치·소견 중심으로 한 칸에 여러 줄.
        (상이처·직전등급·검토사항은 별도 컬럼으로 분리 — 여기 중복 기재 안 함)"""
        L = []
        if ag.get("base_date") or ag.get("grade_date"):
            L.append(f"○ 기준일자 {ag.get('base_date') or '—'} / 등급기준일 {ag.get('grade_date') or '—'}")
        ms = _measure_str(ag)
        if ms:
            L.append(f"○ 신체검사 측정치: {ms}")
        if ag.get("specialist_opinion"):
            L.append(f"○ 전문의 소견: {ag['specialist_opinion']}")
        if ag.get("onset_narrative"):
            L.append(f"○ 발생경위: {ag['onset_narrative']}")
        if ag.get("prior_history"):
            L.append(f"○ 이전 판정·재심의 경위: {ag['prior_history']}")
        if ag.get("past_history"):
            L.append(f"○ 과거력·기왕증: {ag['past_history']}")
        if ag.get("route_note"):
            L.append(f"○ 경로사항: {ag['route_note']}")
        return "\n".join(L) or "—"

    LEFT_KEYS = ("__opinion__", "injury", "review_items", "note_items", "related_docs")
    ri = hr + 1
    for ag in ags:
        # 상이처별 행: 상이처마다 직전등급→신검과목→신검등급→소견→제안등급을 매기고
        # 그중 최고(중한) 등급을 종합 제안등급으로 산출. 인적사항·종합은 세로 병합.
        items = _items(ag)
        preds = [_predict_grade(it.get("injury"), it.get("body_part") or ag.get("body_part"), emb)
                 for it in items]
        props = [_proposed(it, preds[k]) for k, it in enumerate(items)]
        total = _total_grade(props) or \
            ("등급기준미달" if any("미달" in (p or "") for p in props) else None)
        # 0721 회의 ⑧: 종합판정(등급 상향 검토) 대상 = 7급 상이처 3개 이상일 때만
        n7 = sum(1 for p in props if _severity(p) == 7)
        if total and len(items) > 1:
            total = f"{total}\n(종합판정 대상 — 7급 {n7}개)" if n7 >= 3 \
                else f"{total}\n(종합판정 비대상)"
        n = len(items)
        def _bullets(it, key, mark, k):
            """상이처별 목록 칸 — 복수 상이처면 실물 양식처럼 [상이처명] 태그 줄로 시작."""
            xs = it.get(key) or []
            if not xs:
                return "—"
            head = [f"[{it.get('injury')}]"] if (n > 1 and key == "review_items") else []
            return "\n".join(head + [f"{mark} {x}" for x in xs])

        for k, it in enumerate(items):
            row_txts = []  # 행 높이 추정용 (해당 행의 긴 텍스트 칸들)
            for ci, (label, key, _w, scope) in enumerate(COLUMNS, 1):
                if scope == "ag":
                    val = (total or "—") if key == "__grade_total__" else fmt(ag.get(key))
                    if key == "applicant" and ag.get("is_real"):
                        val = f"{val}\n[실데이터]"
                    if k > 0:
                        val = None  # 병합 범위 — 값은 첫 행에만
                elif key == "__opinion__":
                    val = it.get("opinion") or (soken(ag) if k == 0 else "—")
                    row_txts.append((val, 40))
                elif key == "review_items":
                    val = _bullets(it, key, "○", k)
                    row_txts.append((val, 36))
                elif key == "note_items":
                    val = _bullets(it, key, "◇", k)
                    ud = _updown(it.get("prev_grade"), props[k])
                    if ud:  # 실물 양식의 비고 '승급' 표기 — 직전등급 대비 자동 판정
                        val = f"[{ud}]" + ("" if val == "—" else f"\n{val}")
                elif key == "related_docs":
                    val = _bullets(it, key, "·", k)
                elif key == "__grade_each__":
                    val = props[k] or "—"
                    if preds[k] and _severity(preds[k]) != _severity(props[k]):
                        val += f"\n(AI 참고: {preds[k]})"
                else:
                    val = fmt(it.get(key))
                c = ws.cell(ri + k, ci, val)
                c.font = cf; c.border = border
                c.alignment = wrap_l if key in LEFT_KEYS else wrap
            wrapped = max(sum(max(1, len(seg) // w + 1) for seg in str(t).split("\n"))
                          for t, w in row_txts) if row_txts else 3
            ws.row_dimensions[ri + k].height = max(44, min(wrapped, 30) * 14)
        if n > 1:  # 안건 단위 칸 세로 병합 (인적사항·관련자료·검토사항·비고·종합 제안등급)
            for ci, (label, key, _w, scope) in enumerate(COLUMNS, 1):
                if scope == "ag":
                    ws.merge_cells(start_row=ri, start_column=ci, end_row=ri + n - 1, end_column=ci)
        ri += n

    ws.freeze_panes = "A5"

    if ga_id and logs_by.get(ags[0]["ga_id"]):
        ws2 = wb.create_sheet("작업로그")
        heads = ["시각", "단계", "이벤트", "수행자", "상세", "첨부파일", "상태"]
        for ci, h in enumerate(heads, 1):
            c = ws2.cell(1, ci, h); c.font = hf; c.fill = hdr_fill; c.border = border; c.alignment = center
        for ri, lg in enumerate(logs_by[ags[0]["ga_id"]], 2):
            vals = [
                lg["created_at"].strftime("%Y-%m-%d %H:%M") if lg.get("created_at") else "",
                lg.get("step") or "", lg.get("event") or "", lg.get("actor") or "",
                lg.get("detail") or "", lg.get("file_name") or "", lg.get("status") or "",
            ]
            for ci, v in enumerate(vals, 1):
                c = ws2.cell(ri, ci, v); c.font = cf; c.border = border
                c.alignment = Alignment(wrap_text=True, vertical="top")
        for col, w in zip("ABCDEFG", (16, 10, 20, 10, 36, 14, 8)):
            ws2.column_dimensions[col].width = w
        ws2.freeze_panes = "A2"

    # 신체검사 측정치 시트 (단건, measurements 있을 때)
    if ga_id and ags[0].get("measurements"):
        import json as _json
        meas = ags[0]["measurements"]
        if isinstance(meas, str):
            meas = _json.loads(meas)
        ws3 = wb.create_sheet("신체검사 측정치")
        mheads = ["검사항목", "측정값", "단위", "기준", "판정"]
        for ci, h in enumerate(mheads, 1):
            c = ws3.cell(1, ci, h); c.font = hf; c.fill = hdr_fill; c.border = border; c.alignment = center
        for ri, m in enumerate(meas, 2):
            vals = [m.get("name", ""), str(m.get("value", "")), m.get("unit", ""),
                    m.get("ref", ""), m.get("result", "")]
            for ci, v in enumerate(vals, 1):
                c = ws3.cell(ri, ci, v); c.font = cf; c.border = border
                c.alignment = center if ci in (2, 3) else Alignment(wrap_text=True, vertical="center")
        for col, w in zip("ABCDE", (30, 16, 8, 20, 12)):
            ws3.column_dimensions[col].width = w
        ws3.freeze_panes = "A2"

    # 의무기록 시간순 시트 (단건, med_timeline 있을 때)
    if ga_id and ags[0].get("med_timeline"):
        import json as _json2
        tl = ags[0]["med_timeline"]
        if isinstance(tl, str):
            tl = _json2.loads(tl)
        ws4 = wb.create_sheet("의무기록 시간순")
        theads = ["진료일", "의료기관", "기록유형", "진단명", "소견"]
        for ci, h in enumerate(theads, 1):
            c = ws4.cell(1, ci, h); c.font = hf; c.fill = hdr_fill; c.border = border; c.alignment = center
        for ri, r in enumerate(tl, 2):
            vals = [r.get("date", ""), r.get("hospital", ""), r.get("type", ""),
                    r.get("dx", ""), r.get("finding", "")]
            for ci, v in enumerate(vals, 1):
                c = ws4.cell(ri, ci, v); c.font = cf; c.border = border
                c.alignment = Alignment(wrap_text=True, vertical="top")
            ws4.row_dimensions[ri].height = 44
        for col, w in zip("ABCDE", (14, 20, 14, 26, 60)):
            ws4.column_dimensions[col].width = w
        ws4.freeze_panes = "A2"

    if ga_id:
        ag = ags[0]
        safe = (ag.get("agenda_no") or f"ga{ga_id}").replace("/", "_")
        fname = f"상이등급심사표_{safe}_{ag.get('applicant','')}.xlsx"
    else:
        fname = "상이등급심사표_전체.xlsx"
    path = os.path.join(tempfile.gettempdir(), fname)
    wb.save(path)
    return fname, path


def export_batch(ga_ids, emb=None):
    """여러 안건의 심사표를 안건별 xlsx로 만들어 zip으로 일괄 산출.
    (의결서 export_batch와 동일 패턴 — 안건별 부속 시트도 각 파일에 포함)"""
    import zipfile
    from datetime import date

    fname = f"상이등급심사표_일괄_{len(ga_ids)}건_{date.today().strftime('%Y%m%d')}.zip"
    path = os.path.join(tempfile.gettempdir(), fname)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        for gid in ga_ids:
            inner_name, inner_path = export_xlsx(gid, emb)
            zf.write(inner_path, arcname=inner_name)
    return fname, path
